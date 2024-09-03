import openpyxl
import os

# Mapping of source columns (0-based index) to destination columns (0-based index)
column_mapping = {
    0: 0,   # H2 -> A (1st column -> 1st column)
    1: 1,   # H3 -> B (2nd column -> 2nd column)
    1: 6,   # B -> G (2nd column -> 7th column, one column over)
    3: 8,   # D -> J (4th column -> 9th column, one column over)
    4: 9,   # E -> K (5th column -> 10th column, one column over)
    5: 10,  # F -> L (6th column -> 11th column, one column over)
    6: 11,  # G -> M (7th column -> 12th column, one column over)
    7: 12,  # H -> N (8th column -> 13th column, one column over)
    8: 14,  # I -> O (9th column -> 14th column, one column over)
    9: 13,  # J -> P (10th column -> 15th column, one column over)
    10: 15, # K -> AE (11th column -> 29th column, one column over)
    11: 16, # L -> T (12th column -> 16th column, one column over)
    12: 30, # M -> U (13th column -> 17th column, one column over)
    13: 19,  # N -> V (14th column -> 18th column, one column over)
    14: 20,  # O -> U
    15: 21,  # P -> 
    17: 40   # R -> AU    
}

def read_and_map_columns(source_files, destination_file, sheet_name='HeaderforMcCookPreApproval'):
    """
    Read specific columns from multiple source Excel files and map them to the destination Excel file.
    :param source_files: List of paths to the source Excel files.
    :param destination_file: Path to the destination Excel file.
    :param sheet_name: Name of the sheet in the destination file.
    """
    # Load or create the destination workbook and sheet
    if os.path.exists(destination_file):
        destination_wb = openpyxl.load_workbook(destination_file)
        if sheet_name in destination_wb.sheetnames:
            destination_ws = destination_wb[sheet_name]
        else:
            destination_ws = destination_wb.create_sheet(sheet_name)
    else:
        destination_wb = openpyxl.Workbook()
        destination_ws = destination_wb.active
        destination_ws.title = sheet_name

    # Iterate through each source file
    for source_file in source_files:
        # Load the source workbook and sheet
        if os.path.exists(source_file):
            source_wb = openpyxl.load_workbook(source_file)
            source_ws = source_wb.active

            # Extract values from specific cells H2 and H3
            H2_value = source_ws['H2'].value
            H3_value = source_ws['H3'].value

            # Determine the starting row for appending data in the destination file
            start_row = destination_ws.max_row + 1

            # Identify the last valid row in column D with data
            last_valid_row = max((idx for idx, cell in enumerate(source_ws['D'], start=1) if cell.value is not None), default=0)

            # Iterate through rows of the source worksheet, skipping the first 10 rows
            for row_idx in range(10, last_valid_row + 1):
                for source_col_idx, dest_col_idx in column_mapping.items():
                    source_cell = source_ws.cell(row=row_idx, column=source_col_idx + 1)
                    destination_ws.cell(row=start_row, column=dest_col_idx + 1).value = source_cell.value

                # Insert H2 and H3 values into Column A and Column B
                destination_ws.cell(row=start_row, column=0 + 1).value = H2_value  # H2 -> A
                destination_ws.cell(row=start_row, column=1 + 1).value = H3_value  # H3 -> B

                start_row += 1

        else:
            print(f"File not found: {source_file}")

    # Save the destination workbook
    destination_wb.save(destination_file)

if __name__ == "__main__":
    # List of paths to the source Excel files
    source_files = [
        r"C:\Python\Wheeling\DATL 2912.xlsx",
        
    ]

    # Path to the destination Excel file
    destination_file = r"C:\Python\Wheeling\HeaderforMcCookPreApproval_Current.xlsx"

    # Sheet name for the destination file
    sheet_name = 'HeaderforMcCookPreApproval'

    # Call the function to read and map columns
    read_and_map_columns(source_files, destination_file, sheet_name)