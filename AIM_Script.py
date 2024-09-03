import openpyxl
import os
 
# Mapping of source columns to destination columns
column_mapping = {
    'A': 'A', 'B': 'B', 
    'F': 'I', 'G': 'J', 'H': 'K', 'I': 'L', 'J': 'M', 
    'L': 'AC', 'M': 'N', 'O': 'O', 'S': 'S', 
    'T': 'Q', 'U': 'R', 'V': 'AE', 'W': 'V'
}
 
def read_and_map_columns(source_files, destination_file, sheet_name='HeaderForMcCookPreApproval2'):
    """
    Read specific columns from multiple source Excel files and map them to the destination Excel file.
    :param source_files: List of paths to the source Excel files.
    :param destination_file: Path to the destination Excel file.
    :param sheet_name: Name of the sheet in the destination file.
    """
    # Load or create the destination workbook and sheet
    if os.path.exists(destination_file):
        destination_wb = openpyxl.load_workbook(destination_file)
        if sheet_name in destination_wb.sheetnames:
            destination_ws = destination_wb[sheet_name]
        else:
            destination_ws = destination_wb.create_sheet(sheet_name)
    else:
        destination_wb = openpyxl.Workbook()
        destination_ws = destination_wb.active
        destination_ws.title = sheet_name
 
    # Iterate through each source file
    for source_file in source_files:
        # Load the source workbook and sheet
        source_wb = openpyxl.load_workbook(source_file)
        source_ws = source_wb.active
 
        # Determine the starting row for appending data in the destination file
        start_row = destination_ws.max_row + 1
 
        # Iterate through rows of the source worksheet (skipping the first row)
        for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, max_row=source_ws.max_row), start=start_row):
            for source_col, dest_col in column_mapping.items():
                source_cell_value = row[openpyxl.utils.column_index_from_string(source_col) - 1].value
                destination_ws[f"{dest_col}{row_idx}"] = source_cell_value
 
    # Save the destination workbook
    destination_wb.save(destination_file)
 
if __name__ == "__main__":
    # List of paths to the source Excel files
    source_files = [
        r"C:\Python\New Invoices\MLCo_Invoice_I022671_WE_01-02-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I022754_WE_01-09-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I022831_WE_01-16-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I022884_WE_01-23-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I022958A_WE_01-30-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023050_WE_02-06-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023099_WE_02-13-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023184_WE_02-20-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023277_WE_02-27-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023327_WE_03-06-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023443_WE_03-13-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023499_WE_03-20-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023613_WE_03-27-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023635_WE_04-03-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023760_WE_04-10-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023848_WE_04-17-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023931_WE_04-24-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I023968_WE_05-01-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024076_WE_05-08-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024158_WE_05-15-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024268_WE_05-22-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024375_WE_05-29-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024391_WE_06-05-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024544_WE_06-12-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024588_WE_06-19-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024717_WE_06-26-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024829_WE_07-03-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024893_WE_07-10-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I024965_WE_07-17-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025045_WE_07-24-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025165_WE_07-31-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025246_WE_08-07-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025300_WE_08-14-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025376_WE_08-21-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025463_WE_08-28-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025544_WE_09-04-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025608_WE_09-11-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025687_WE_09-18-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025745_WE_09-25-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025810_WE_10-02-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025867_WE_10-09-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I025976_WE_10-16-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026049_WE_10-23-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026150_WE_10-30-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026208_WE_11-06-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026251_WE_11-13-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026356_WE_11-20-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026453_WE_11-27-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026487_WE_12-04-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026581_WE_12-11-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026676_WE_12-18-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026750_WE_12-25-21.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026823_WE_01-01-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026896_WE_01-08-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I026951_WE_01-15-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027051_WE_01-22-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027086_WE_01-29-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027148_WE_02-05-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027235_WE_02-12-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027313_WE_02-19-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027374_WE_02-26-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027453_WE_03-05-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027558_WE_03-12-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027608_WE_03-19-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027708_WE_03-26-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027750_WE_04-02-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027866_WE_04-09-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I027922_WE_04-16-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028026_WE_04-23-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028084_WE_04-30-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028159_WE_05-07-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028342_WE_05-28-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028476_WE_06-04-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028528_WE_06-11-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028612_WE_06-18-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028699_WE_06-25-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I028770_WE_07-02-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I030155_WE_11-05-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I030206_WE_11-12-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I030275_WE_11-19-22.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I030999_WE_01-28-23.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I031070_WE_02-04-23.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I031168_WE_02-11-23.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I031282_WE_02-18-23.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I031359_WE_02-25-23.xlsx",
        r"C:\Python\New Invoices\MLCo_Invoice_I031409_WE_03-04-23.xlsx"
    ]
 
    # Path to the destination Excel file
    destination_file = r"C:\Python\New Invoices\HeaderforMcCookPreApproval2.xlsx"
 
    read_and_map_columns(source_files, destination_file)